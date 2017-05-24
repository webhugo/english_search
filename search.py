#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Created by webhugo on 17-5-14.


import os
import sys
import config

from openpyxl import load_workbook

# 方法二
# import urllib2
# from bs4 import BeautifulSoup


global list
list = []
strl = "curl -s http://dict.cn/"
strr = """ | grep "<li><strong>" | tr -d '\t' | sed 's/<li><strong>//g' | sed 's/<\/li>//g' | sed 's/<\/strong>//g'
"""


def _get_index_file(char):
    for file in config.files:
        if char == file[file.rindex("/") + 1:file.rindex(config.postfix)]:
            return file


def _spider(word, file):
    # 方法一
    str = strl + word + strr
    str = os.popen(str)
    str = str.read()
    sys.stdout.flush()
    str = str.split('\n')
    wb = load_workbook(file)
    sheetNames = wb.sheetnames
    ws = wb[sheetNames[0]]
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=word)
    i = 2
    for s in str:
        if not len(s.strip()) == 0:
            list.append(s)
            ws.cell(row=row, column=i, value=s)
        i += 1
    wb.save(file)


    # with open(file, 'a') as  file:
    #     flag = True  # 表示第一次写入
    #     for s in str:
    #         if not len(s) == 0:
    #             list.append(s)
    #             if flag == True:
    #                 file.write("\n")
    #                 file.write("|\t" + word + "\t|\t" + s + "\t|")
    #                 flag = False
    #             else:
    #                 file.write("\t" + s + "\t|")

    # 方法二

    # str = []
    # content = urllib2.urlopen("http://dict.cn/"+word).read()
    # soup = BeautifulSoup(content, 'html.parser')
    # div = soup.find_all("div", class_="layout dual")
    # for d in div:
    #     for li in d.find_all("li"):
    #         str.append(li.text)
    # with open(file, 'a') as  file:
    #     flag = True  # 表示第一次写入
    #     for s in str:
    #         if not len(s) == 0:
    #             list.append(s)
    #             if flag == True:
    #                 file.write("\n")
    #                 file.write("|\t" + word + "\t|\t" + s.encode("utf-8") + "\t|")
    #                 flag = False
    #             else:
    #                 file.write("\t" + s.encode("utf-8") + "\t|")


def _search_local_file(word, file):
    wb = load_workbook(file, read_only=True)
    sheetNames = wb.sheetnames
    ws = wb[sheetNames[0]]
    for row in ws.rows:
        try:
            if row[0].value == word:
                for cell in row:
                    list.append(cell.value)
                list.pop(0)
        except:
            pass

    if not len(list) == 0:
        return True
    else:
        return False


def search(word):
    # 删除旧的记录
    del list[:]
    # 获取搜索的文件名字
    file = _get_index_file(word[0])
    if file == None:
        print "no real word"
        return list
    # 首先查找本地文件中是否存在，减少网络请求
    if not _search_local_file(word, file):
        # 如果不存在，则通过爬虫获取
        _spider(word, file)
    if len(list) == 0:
        print "no real word"
        return list

    return list
